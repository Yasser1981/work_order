# -*- coding: utf-8 -*-
"""تصدير أمر العمل إلى ملف إكسل قابل للتعديل (ق-٥٧، وأُعيد تشكيله في ق-٧١).

**الغرض بنصّ المستخدم:** «في بعض الحالات يتطلّب التعديل والحساب يدوياً بدون
الالتزام بحسابات البرنامج».

فهذا القالب **ليس صورة للنتيجة بل ورقة عمل حيّة**: الكميات والأسعار تُكتب
**أرقاماً**، والكلفة **معادلة** — فتعديل أي كمية يُحدِّث كلفتها ومجموعها داخل
الإكسل بلا عودة إلى البرنامج. وهو **الوجهة الوحيدة التي تخرج فيها الأرقام من
سيطرة المحرك** عمداً.

## الشكل (ق-٧١)

الورقة الأولى تحاكي نموذج الإيزو المطبوع بعد تعديلات ق-٦٤ … ق-٦٩:

```
        الترويسة الرسمية والعنوان
        ┌ الترويسة: التبويب · التاريخ · المشروع · المدة · الكلفة … ┐
        ├── الجداول الجانبية ─┬─ أ - جدول المواد المخمنة ──────────┤
        │  ب - الاشراف الفني  │  (بعمودَي السعر والكلفة زيادةً)     │
        │  ج - الاليات        │                                    │
        │  ملاحظات إضافية     │                                    │
        ├──────────── التواقيع الخمسة ───────────────────────────┤
```

**وفرقان مقصودان عن المطبوع:**

1. **جدول المواد هنا يحمل «سعر الوحدة» و«الكلفة»** — والنموذج الرسمي جدول
   كميات لا كلف (ق-٥٧). ولولاهما لضاع سبب وجود هذا الملف أصلاً.
2. **ورقة ثانية «أجور العمل»** لا نظير لها في المطبوع: النموذج الرسمي يحمل
   الكلفة رقماً واحداً في الترويسة. وسطر الكلفة في الورقة الأولى **يشير
   إليها بمعادلة**، فتعديل أجرٍ فيها يُحدّث الترويسة.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from engine.workorder import WorkOrder

from .iso_form import SIGNATURES

_HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
_HEAD_FONT = Font(bold=True, color="FFFFFF")
_TOTAL_FILL = PatternFill("solid", fgColor="F0F0F0")
_THIN = Side(style="thin", color="BBBBBB")
_BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_MONEY = "#,##0"

_RIGHT = Alignment(horizontal="right", vertical="top", wrap_text=True)
_CENTRE = Alignment(horizontal="center", vertical="center", wrap_text=True)

LABOUR_SHEET = "أجور العمل"
ORDER_SHEET = "أمر العمل"

MATERIAL_COLUMNS = ["ت", "اسم المادة", "الوحدة القياسية", "الكمية",
                    "سعر الوحدة", "الكلفة"]
STAFF_COLUMNS = ["ت", "نوع العاملين", "العدد", "عدد الأيام"]
EQUIPMENT_COLUMNS = ["ت", "نوع الآلية", "الرقم", "عدد الأيام"]

SIDE = 8
"""أول عمود في الكتلة الجانبية (H). والعمود G فاصل — نظير خلية الفصل في ق-٦٩."""


def _sheet(workbook: Workbook, title: str):
    sheet = workbook.create_sheet(title)
    # الورقة عربية الاتجاه: إكسل يقلب الأعمدة بنفسه، فالعمود A يظهر **أقصى
    # اليمين** — وهو ما يجعل «ت» يمينَ الجدول كما في المطبوع (ق-٦٤)
    sheet.sheet_view.rightToLeft = True
    return sheet


def _widths(sheet, widths: dict[int, int]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width


def _table_head(sheet, row: int, first_column: int, labels: list[str]) -> None:
    for offset, label in enumerate(labels):
        cell = sheet.cell(row, first_column + offset, label)
        cell.font, cell.fill = _HEAD_FONT, _HEAD_FILL
        cell.alignment = _CENTRE
        cell.border = _BOX


def _boxed(sheet, row: int, first_column: int, count: int) -> None:
    for column in range(first_column, first_column + count):
        sheet.cell(row, column).border = _BOX


def _title(sheet, row: int, column: int, text: str) -> None:
    cell = sheet.cell(row, column, text)
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="right")


# ──────────────────────────── الترويسة ────────────────────────────


def _banner(sheet, order: WorkOrder, last_column: int) -> int:
    """اسم الشركة والفرع ورمز النموذج ورقم أمر العمل. يعيد الصفّ التالي."""
    span = f"A{{row}}:{get_column_letter(last_column)}{{row}}"
    for row, (text, size, align) in enumerate((
        (order.organisation, 14, "center"),
        (order.branch, 11, "center"),
        (order.form_code, 9, "left"),
        (f"أمر عمل رقم {order.number}", 15, "center"),
    ), start=1):
        sheet.merge_cells(span.format(row=row))
        cell = sheet.cell(row, 1, text)
        cell.font = Font(bold=size >= 14, size=size)
        cell.alignment = Alignment(horizontal=align)
    return 6                      # الصفّ 5 فاصل (ق-٦٩)


def _cost_formula(materials_total_row: int, labour_total_row: int) -> str:
    """سطر الكلفة **معادلةً حيّة** لا نصّاً جامداً (ق-٦٥، ق-٧١).

    يحاكي نصّ المطبوع حرفاً بحرف، لكنه يقرأ الرقمين من الورقتين — فتعديل كمية
    مادة أو أجر فقرة يُحدّث الترويسة نفسها. ولو كُتب نصّاً لبقي رقم الترويسة
    على حاله بعد أول تعديل، **وهو أخطر ما يمكن أن يحمله هذا الملف**: ورقة
    تُظهر مجموعاً لا يطابق سطورها.
    """
    materials = f"F{materials_total_row}"
    labour = f"'{LABOUR_SHEET}'!G{labour_total_row}"
    return (
        f'="كلفة المواد التخمينية "&TEXT({materials},"#,##0")'
        f'&" + كلفة العمل التخمينية "&TEXT({labour},"#,##0")'
        f'&" = الكلفة التخمينية الكلية "&TEXT({materials}+{labour},"#,##0")&" د.ع"'
    )


def _header_block(sheet, order: WorkOrder, row: int, cost: str) -> int:
    """كتلة الترويسة بترتيب المطبوع نفسه. تعيد الصفّ التالي."""
    def label(at: int, column: int, text: str) -> None:
        cell = sheet.cell(at, column, text)
        cell.font, cell.fill, cell.border = Font(bold=True), _TOTAL_FILL, _BOX
        cell.alignment = Alignment(horizontal="right", vertical="center")

    def value(at: int, first: int, last: int, text) -> None:
        sheet.merge_cells(start_row=at, start_column=first, end_row=at, end_column=last)
        cell = sheet.cell(at, first, text)
        cell.alignment = _RIGHT
        for column in range(first, last + 1):
            sheet.cell(at, column).border = _BOX

    date = order.order_date.strftime("%Y/%m/%d") if order.order_date else ""
    start = order.start_date.strftime("%Y/%m/%d") if order.start_date else ""

    label(row, 1, "التبويب")
    value(row, 2, 3, order.classification)
    label(row, 4, "التاريخ")
    value(row, 5, 6, date)

    for offset, (text, content) in enumerate((
        ("اسم المشروع وموقعه", order.project_name),
        ("المدة اللازمة لتنفيذ العمل", order.duration),
        ("الكلفة التخمينية للمواد + العمل", cost),
        ("حجم العمل المخطط تنفيذه", order.work_scope),
        ("تاريخ المباشرة بالعمل", start),
    ), start=1):
        label(row + offset, 1, text)
        value(row + offset, 2, 6, content)
    return row + 7                # صفّ فاصل بعد الكتلة (ق-٦٩)


# ──────────────────────────── الجداول ────────────────────────────


def _materials_block(sheet, result: dict, row: int) -> tuple[int, int]:
    """جدول المواد بمعادلاته. يعيد (صفّ المجموع، الصفّ التالي)."""
    _title(sheet, row, 1, "أ - جدول المواد المخمنة")
    _table_head(sheet, row + 1, 1, MATERIAL_COLUMNS)

    first = row + 2
    at = first
    materials = [m for m in result["المواد"] if m["الكمية"] > 0]
    for index, material in enumerate(materials, start=1):
        sheet.cell(at, 1, index).alignment = _CENTRE
        sheet.cell(at, 2, material["المادة"]).alignment = _RIGHT
        sheet.cell(at, 3, material["الوحدة"]).alignment = _CENTRE
        sheet.cell(at, 4, material["الكمية"]).alignment = _CENTRE
        price = 0 if material["كمية_فقط"] else (material["سعر الوحدة"] or 0)
        sheet.cell(at, 5, price).number_format = _MONEY
        # **معادلة لا رقماً** — تعديل الكمية أو السعر يُحدّث الكلفة والمجموع
        sheet.cell(at, 6, f"=D{at}*E{at}").number_format = _MONEY
        _boxed(sheet, at, 1, len(MATERIAL_COLUMNS))
        at += 1

    total_row = at
    _total(sheet, total_row, span=5, label="مجموع كلفة المواد",
           formula=f"=SUM(F{first}:F{at - 1})" if at > first else 0)
    return total_row, total_row + 2


def _people_block(sheet, row: int, title: str, columns: list[str],
                  entries: list, name_of) -> int:
    """جدول جانبي (الإشراف أو الآليات). يعيد الصفّ التالي.

    **بلا أسطر صفرية** كالمطبوع (ق-٦٤): يُطبع المُدخَل وحده، وحين لا مُدخَل
    يبقى صفٌّ فارغ واحد فلا يظهر رأسُ جدول بلا جسم.
    """
    _title(sheet, row, SIDE, title)
    _table_head(sheet, row + 1, SIDE, columns)

    at = row + 2
    used = [entry for entry in entries if entry.count]
    for index, entry in enumerate(used, start=1):
        sheet.cell(at, SIDE, index).alignment = _CENTRE
        sheet.cell(at, SIDE + 1, name_of(entry)).alignment = _RIGHT
        sheet.cell(at, SIDE + 2, entry.count).alignment = _CENTRE
        sheet.cell(at, SIDE + 3, entry.days).alignment = _CENTRE
        _boxed(sheet, at, SIDE, len(columns))
        at += 1
    if not used:
        _boxed(sheet, at, SIDE, len(columns))
        at += 1
    return at + 1                 # صفّ فاصل


def _notes_block(sheet, order: WorkOrder, row: int) -> int:
    _title(sheet, row, SIDE, "ملاحظات إضافية")
    last = SIDE + len(STAFF_COLUMNS) - 1
    sheet.merge_cells(start_row=row + 1, start_column=SIDE,
                      end_row=row + 2, end_column=last)
    sheet.cell(row + 1, SIDE, order.notes).alignment = _RIGHT
    for at in (row + 1, row + 2):
        _boxed(sheet, at, SIDE, len(STAFF_COLUMNS))
    return row + 4


def _signatures(sheet, row: int) -> None:
    """التواقيع الخمسة بترتيبها من اليمين (ق-٦٦)."""
    for offset, name in enumerate(SIGNATURES):
        column = 1 + offset * 2
        sheet.merge_cells(start_row=row, start_column=column,
                          end_row=row, end_column=column + 1)
        cell = sheet.cell(row, column, name)
        cell.font, cell.alignment = Font(bold=True), _CENTRE
        sheet.merge_cells(start_row=row + 1, start_column=column,
                          end_row=row + 1, end_column=column + 1)
        sheet.cell(row + 1, column, "................").alignment = _CENTRE


# ──────────────────────────── البناء ────────────────────────────


def build_workbook(order: WorkOrder, result: dict) -> Workbook:
    """ورقة تحاكي النموذج المطبوع، وورقة لأجور العمل بمعادلاتها (ق-٧١)."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    sheet = _sheet(workbook, ORDER_SHEET)
    labour = _sheet(workbook, LABOUR_SHEET)

    labour_total_row = _build_labour_sheet(labour, result)
    _build_order_sheet(sheet, order, result, labour_total_row)
    return workbook


def _build_order_sheet(sheet, order: WorkOrder, result: dict,
                       labour_total_row: int) -> None:
    _widths(sheet, {1: 6, 2: 44, 3: 14, 4: 12, 5: 16, 6: 18,
                    7: 3, 8: 6, 9: 26, 10: 10, 11: 12})

    row = _banner(sheet, order, last_column=SIDE + len(STAFF_COLUMNS) - 1)

    # عدد المواد يحدّد صفّ المجموع، وسطر الكلفة يشير إليه — فيُحسب أولاً
    counted = len([m for m in result["المواد"] if m["الكمية"] > 0])
    header_rows = 6 + 1                       # ستّة صفوف وفاصل
    materials_total_row = row + header_rows + 2 + counted

    row = _header_block(sheet, order, row,
                        _cost_formula(materials_total_row, labour_total_row))

    side = row
    total_row, after_materials = _materials_block(sheet, result, row)
    assert total_row == materials_total_row, (total_row, materials_total_row)

    side = _people_block(sheet, side, "ب - الاشراف الفني", STAFF_COLUMNS,
                         order.staff, lambda entry: entry.role)
    side = _people_block(sheet, side, "ج - الاليات والمعدات", EQUIPMENT_COLUMNS,
                         order.equipment, lambda entry: entry.name)
    side = _notes_block(sheet, order, side)

    _signatures(sheet, max(after_materials, side) + 1)
    sheet.sheet_view.showGridLines = False


def _build_labour_sheet(sheet, result: dict) -> int:
    """ورقة الأجور بمعادلاتها. تعيد صفّ المجموع ليشير إليه سطر الكلفة."""
    _widths(sheet, {1: 6, 2: 18, 3: 46, 4: 12, 5: 14, 6: 16, 7: 18})
    _table_head(sheet, 1, 1,
                ["ت", "الباب", "الفقرة", "الكمية", "الوحدة", "السعر", "الكلفة"])

    row = 2
    for index, line in enumerate(result["أجور_العمل"], start=1):
        sheet.cell(row, 1, index).alignment = _CENTRE
        sheet.cell(row, 2, line.group or "الأعمال الكهربائية").alignment = _RIGHT
        sheet.cell(row, 3, line.name).alignment = _RIGHT
        sheet.cell(row, 4, line.qty).alignment = _CENTRE
        sheet.cell(row, 5, line.unit).alignment = _CENTRE
        sheet.cell(row, 6, line.rate or 0).number_format = _MONEY
        sheet.cell(row, 7, f"=D{row}*F{row}").number_format = _MONEY
        _boxed(sheet, row, 1, 7)
        row += 1

    _total(sheet, row, span=6, label="مجموع أجور التنفيذ",
           formula=f"=SUM(G2:G{row - 1})" if row > 2 else 0)
    sheet.freeze_panes = "A2"
    return row


def _total(sheet, row: int, span: int, label: str, formula) -> None:
    cell = sheet.cell(row, 1, label)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="right")
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    total = sheet.cell(row, span + 1, formula)
    total.font = Font(bold=True)
    total.number_format = _MONEY
    for column in range(1, span + 2):
        sheet.cell(row, column).fill = _TOTAL_FILL
        sheet.cell(row, column).border = _BOX


def write_xlsx(order: WorkOrder, result: dict, path: str) -> str:
    """يكتب المصنَّف إلى المسار ويعيده."""
    if not path.lower().endswith(".xlsx"):
        path += ".xlsx"
    build_workbook(order, result).save(path)
    return path
